import io
from datetime import datetime
from sqlalchemy.orm import Session
from sqlalchemy import text
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def generate_executive_summary_xlsx(db: Session):
    """
    Generates a production-ready Executive Summary Excel report for LogVision.
    Queries PostgreSQL base_event for log volumes and ml_sequence_score for ML insights.
    """
    # 1. Configuration & Constants
    ACTIVE_MODEL = "logbert_like_distilbert_iforest"
    ACTIVE_VERSION = "logbert_v1_full"
    
    # 2. Data Extraction
    # Total log volume and ingestion stats
    stats_query = text("""
        SELECT 
            count(*) as total_logs,
            count(*) FILTER (WHERE log_level IN ('ERROR', 'CRITICAL', 'FATAL')) as error_count,
            MAX(event_timestamp) as last_event
        FROM base_event
    """)
    stats = db.execute(stats_query).fetchone()
    
    # Anomaly stats for the active model
    anomaly_stats_query = text("""
        SELECT count(*) as total_anomalies
        FROM ml_sequence_score 
        WHERE model_name = :m AND model_version = :v AND final_anomaly_label = 'anomaly'
    """)
    anomalies = db.execute(anomaly_stats_query, {"m": ACTIVE_MODEL, "v": ACTIVE_VERSION}).fetchone()

    # Detailed Anomalies List
    details_query = text("""
        SELECT sequence_uid, application_key, component_name, final_anomaly_score, created_at
        FROM ml_sequence_score
        WHERE model_name = :m AND model_version = :v
        ORDER BY final_anomaly_score DESC
        LIMIT 1000
    """)
    details = db.execute(details_query, {"m": ACTIVE_MODEL, "v": ACTIVE_VERSION}).fetchall()

    # 3. Excel Generation with openpyxl
    wb = Workbook()
    
    # --- TAB 1: EXECUTIVE OVERVIEW ---
    ws_summary = wb.active
    ws_summary.title = "Executive Overview"
    
    # Header Styling
    title_font = Font(bold=True, size=16, color="1F4E78")
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    bold_font = Font(bold=True)

    ws_summary.merge_cells('A1:E1')
    ws_summary['A1'] = "LogVision Executive Summary"
    ws_summary['A1'].font = title_font
    ws_summary['A1'].alignment = Alignment(horizontal='center')

    ws_summary['A3'] = "Report Timestamp:"
    ws_summary['B3'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws_summary['A4'] = "Active ML Model:"
    ws_summary['B4'] = f"{ACTIVE_MODEL} ({ACTIVE_VERSION})"

    # KPI Table
    ws_summary.append([]) # Spacer
    ws_summary.append(["Metric", "Value", "Notes"])
    for cell in ws_summary[6]:
        cell.font = bold_font
        cell.fill = header_fill

    kpis = [
        ["Total Logs Analyzed", stats.total_logs, "Source: base_event"],
        ["Anomalies Detected", anomalies.total_anomalies, "Label: anomaly"],
        ["Critical Error Volume", stats.error_count, "Levels: ERROR, CRITICAL, FATAL"],
        ["Stability Index", f"{round((1 - (anomalies.total_anomalies/max(stats.total_logs, 1)))*100, 2)}%", "Derived score"]
    ]
    for kpi in kpis:
        ws_summary.append(kpi)

    # --- TAB 2: ANOMALY DETAILS ---
    ws_details = wb.create_sheet("Anomaly Details")
    headers = ["Sequence UID", "Application", "Component", "Anomaly Score", "Detected At"]
    ws_details.append(headers)
    for cell in ws_details[1]:
        cell.font = bold_font
        cell.fill = header_fill

    for row in details:
        ws_details.append([
            row.sequence_uid,
            row.application_key,
            row.component_name,
            float(row.final_anomaly_score),
            row.created_at.replace(tzinfo=None) if row.created_at else None
        ])

    # Auto-adjust column widths for readability
    for ws in wb.worksheets:
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except: pass
            ws.column_dimensions[column].width = max_length + 2

    # 4. Return as Byte Stream for FastAPI
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    return file_stream